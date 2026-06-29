"""Excel report generation."""

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .logger import ApplicationLogger
from .models import MatchResult


class Reporter:
    """Generates Excel reports."""

    def __init__(self, logger: Optional[ApplicationLogger] = None) -> None:
        """Initialize the reporter.

        Args:
            logger: Application logger instance.
        """
        self.logger = logger or ApplicationLogger()

    def generate_results_report(
        self, match_results: list[MatchResult], output_path: Path
    ) -> bool:
        """Generate main results report.

        Args:
            match_results: List of match results.
            output_path: Path for the Excel file.

        Returns:
            True if successful, False otherwise.
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados"

            # Header
            headers = ["Data", "Paciente", "Procedimento", "Pago", "Arquivo", "Página"]
            ws.append(headers)

            # Style header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data rows
            for result in match_results:
                ws.append(
                    [
                        result.agenda_entry.data.strftime("%d/%m/%Y %H:%M"),
                        result.agenda_entry.paciente,
                        result.agenda_entry.procedimento,
                        "Sim" if result.is_paid else "Não",
                        result.agenda_entry.arquivo,
                        result.agenda_entry.pagina,
                    ]
                )

            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

            # Save
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            self.logger.info(f"Results report saved: {output_path}")
            return True

        except Exception as e:
            self.logger.error(f"Error generating results report: {str(e)}")
            return False

    def generate_not_found_report(
        self, not_found_payments: list, output_path: Path
    ) -> bool:
        """Generate report of payments not found in agenda.

        Args:
            not_found_payments: List of PaymentEntry objects not matched.
            output_path: Path for the Excel file.

        Returns:
            True if successful, False otherwise.
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Não Encontrados"

            # Header
            headers = ["Data", "Paciente", "Procedimento", "Valor"]
            ws.append(headers)

            # Style header
            header_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data rows
            for payment in not_found_payments:
                ws.append(
                    [
                        payment.data.strftime("%d/%m/%Y"),
                        payment.paciente,
                        payment.procedimento,
                        f"R$ {payment.valor:.2f}",
                    ]
                )

            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

            # Save
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            self.logger.info(f"Not found report saved: {output_path}")
            return True

        except Exception as e:
            self.logger.error(f"Error generating not found report: {str(e)}")
            return False
